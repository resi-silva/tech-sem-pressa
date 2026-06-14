import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: SEMANA (planejamento semanal)
# ============================================================
ws = wb.active
ws.title = "Planejamento Semanal"

# Cores
cor_cabecalho = PatternFill(start_color="7D9B7A", fill_type="solid")
cor_subcab = PatternFill(start_color="A8C3A6", fill_type="solid")
cor_destaque = PatternFill(start_color="D4C9E0", fill_type="solid")
cor_linha_alt = PatternFill(start_color="FDFCFA", fill_type="solid")
cor_branco = PatternFill(start_color="FFFFFF", fill_type="solid")
cor_legenda = PatternFill(start_color="F0EBE0", fill_type="solid")
cor_obs = PatternFill(start_color="F5F0EB", fill_type="solid")
cor_recados = PatternFill(start_color="E8F0E6", fill_type="solid")
cor_imprev = PatternFill(start_color="C9B38B", fill_type="solid")

fonte_titulo_play = Font(name="Playfair Display", size=18, bold=True, color="2D2D2D")
fonte_subtitulo = Font(name="Playfair Display", size=13, bold=True, color="2D2D2D")
fonte_cabecalho = Font(name="Inter", size=11, bold=True, color="FFFFFF")
fonte_subcab = Font(name="Inter", size=10, bold=True, color="2D2D2D")
fonte_normal = Font(name="Inter", size=10, color="3A3A3A")
fonte_destaque = Font(name="Inter", size=10, bold=True, color="5A7A57")
fonte_peq = Font(name="Inter", size=9, color="6B6B6B")

alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
alinhamento_esq = Alignment(horizontal="left", vertical="center", wrap_text=True)
alinhamento_topo = Alignment(horizontal="left", vertical="top", wrap_text=True)

borda_fina = Border(
    left=Side(style="thin", color="E0DCD4"),
    right=Side(style="thin", color="E0DCD4"),
    top=Side(style="thin", color="E0DCD4"),
    bottom=Side(style="thin", color="E0DCD4")
)

# --- TITULO PRINCIPAL ---
ws.merge_cells("A1:H1")
ws["A1"] = "\U0001f5d3\ufe0f  PLANEJAMENTO SEMANAL \u2014 M\u00c3E DE AUTISTA"
ws["A1"].font = fonte_titulo_play
ws["A1"].alignment = alinhamento_centro
ws.row_dimensions[1].height = 45

# --- SUBTITULO ---
ws.merge_cells("A2:H2")
ws["A2"] = "Uma semana organizada \u00e9 uma semana com mais acolhimento, menos sobrecarga e mais tempo para o que realmente importa."
ws["A2"].font = fonte_peq
ws["A2"].alignment = alinhamento_centro
ws.row_dimensions[2].height = 28

# --- CABECALHO DA SEMANA ---
ws.merge_cells("A4:H4")
ws["A4"] = "SEMANA DE: ____________________  |  PER\u00cdODO: ____/____ A ____/____"
ws["A4"].font = Font(name="Inter", size=12, bold=True, color="2D2D2D")
ws["A4"].alignment = alinhamento_centro
ws["A4"].fill = cor_legenda
ws.row_dimensions[4].height = 32

# --- LEGENDA ---
ws.merge_cells("A5:H5")
ws["A5"] = "LEGENDA:  Rotina Filho(a)  |  Minhas Tarefas  |  Cuidados Pessoais  |  Consultas/Terapias  |  Lazer  |  Imprevistos"
ws["A5"].font = fonte_peq
ws["A5"].alignment = alinhamento_centro
ws.row_dimensions[5].height = 22

# --- COLUNAS ---
colunas = ["Hor\u00e1rio", "Segunda", "Ter\u00e7a", "Quarta", "Quinta", "Sexta", "S\u00e1bado", "Domingo"]
for i, col in enumerate(colunas, 1):
    cell = ws.cell(row=7, column=i, value=col)
    cell.font = fonte_cabecalho
    cell.fill = cor_cabecalho
    cell.alignment = alinhamento_centro
    cell.border = borda_fina
ws.row_dimensions[7].height = 30

# --- LINHAS DE HORARIOS ---
horarios = [
    "06h - 08h  |  Despertar / Rotina Matinal",
    "08h - 10h  |  Estudos / Terapias",
    "10h - 12h  |  Atividades / Trabalho",
    "12h - 14h  |  Almo\u00e7o / Descanso",
    "14h - 16h  |  Terapias / Tarefas",
    "16h - 18h  |  Livre / Lazer / Estimula\u00e7\u00e3o",
    "18h - 20h  |  Jantar / Rotina Noturna",
    "20h - 22h  |  Planejamento / Autocuidado",
    "22h +      |  Dormir / Descanso"
]

for idx, horario in enumerate(horarios):
    row = 8 + idx
    ws.row_dimensions[row].height = 50
    cell_h = ws.cell(row=row, column=1, value=horario)
    cell_h.font = fonte_subcab
    cell_h.fill = cor_subcab
    cell_h.alignment = alinhamento_centro
    cell_h.border = borda_fina

    for col in range(2, 9):
        cell = ws.cell(row=row, column=col)
        cell.fill = cor_branco if idx % 2 == 0 else cor_linha_alt
        cell.border = borda_fina
        cell.alignment = alinhamento_topo

# --- LINHA DE OBSERVACOES ---
row_obs = 8 + len(horarios)
ws.merge_cells(f"A{row_obs}:H{row_obs}")
ws[f"A{row_obs}"] = "OBSERVA\u00c7\u00d5ES DA SEMANA:  __________________________________________________________________"
ws[f"A{row_obs}"].font = fonte_destaque
ws[f"A{row_obs}"].alignment = alinhamento_esq
ws[f"A{row_obs}"].fill = cor_obs
ws.row_dimensions[row_obs].height = 30

# --- CHECKLIST DIARIO ---
row_check = row_obs + 2
ws.merge_cells(f"A{row_check}:H{row_check}")
ws[f"A{row_check}"] = "CHECKLIST DI\u00c1RIO \u2014 O QUE N\u00c3O PODE FALTAR"
ws[f"A{row_check}"].font = fonte_subtitulo
ws[f"A{row_check}"].alignment = alinhamento_esq
ws.row_dimensions[row_check].height = 30

itens_check = [
    "Rotina de sono respeitada",
    "Medica\u00e7\u00e3o administrada (se houver)",
    "Refei\u00e7\u00f5es equilibradas",
    "Atividade f\u00edsica / estimula\u00e7\u00e3o",
    "Tempo de tela controlado",
    "Momento de conex\u00e3o (abra\u00e7o, brincadeira, conversa)",
    "Paci\u00eancia renovada (respirei fundo antes de reagir)",
    "Um momento s\u00f3 meu (5 min que seja)",
]

for i, item in enumerate(itens_check):
    r = row_check + 1 + i
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = f"  \u2610  {item}"
    ws[f"A{r}"].font = fonte_normal
    ws[f"A{r}"].alignment = alinhamento_esq
    ws[f"A{r}"].fill = cor_linha_alt if i % 2 == 0 else cor_branco
    ws.row_dimensions[r].height = 24

# --- METAS E CONQUISTAS ---
row_metas = row_check + 1 + len(itens_check) + 2
ws.merge_cells(f"A{row_metas}:D{row_metas}")
ws[f"A{row_metas}"] = "METAS DA SEMANA"
ws[f"A{row_metas}"].font = fonte_subtitulo
ws[f"A{row_metas}"].alignment = alinhamento_esq
ws.row_dimensions[row_metas].height = 30

ws.merge_cells(f"E{row_metas}:H{row_metas}")
ws[f"E{row_metas}"] = "CONQUISTAS / APRENDIZADOS"
ws[f"E{row_metas}"].font = fonte_subtitulo
ws[f"E{row_metas}"].alignment = alinhamento_esq
ws.row_dimensions[row_metas].height = 30

for i in range(6):
    r = row_metas + 1 + i
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = f"  {i+1}.  ___________________________________________"
    ws[f"A{r}"].font = fonte_normal
    ws[f"A{r}"].alignment = alinhamento_esq
    ws.row_dimensions[r].height = 24

    ws.merge_cells(f"E{r}:H{r}")
    ws[f"E{r}"] = f"  {i+1}.  ___________________________________________"
    ws[f"E{r}"].font = fonte_normal
    ws[f"E{r}"].alignment = alinhamento_esq
    ws.row_dimensions[r].height = 24

# --- RECADOS ---
row_recados = row_metas + 1 + 6 + 1
ws.merge_cells(f"A{row_recados}:H{row_recados}")
ws[f"A{row_recados}"] = "RECADOS / LEMBRETES  |  __________________________________________________________________"
ws[f"A{row_recados}"].font = fonte_destaque
ws[f"A{row_recados}"].alignment = alinhamento_esq
ws[f"A{row_recados}"].fill = cor_recados
ws.row_dimensions[row_recados].height = 28

# --- LARGURA DAS COLUNAS ---
ws.column_dimensions["A"].width = 24
for col in range(2, 9):
    ws.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# SHEET 2: TERAPIAS E PROFISSIONAIS
# ============================================================
ws2 = wb.create_sheet("Terapias e Profissionais")

ws2.merge_cells("A1:F1")
ws2["A1"] = "TERAPIAS, PROFISSIONAIS E ACOMPANHAMENTOS"
ws2["A1"].font = fonte_titulo_play
ws2["A1"].alignment = alinhamento_centro
ws2.row_dimensions[1].height = 40

cabecalho2 = ["Profissional", "Especialidade", "Dias / Hor\u00e1rios", "Contato", "Pr\u00f3xima Consulta", "Observa\u00e7\u00f5es"]
for i, col in enumerate(cabecalho2, 1):
    cell = ws2.cell(row=3, column=i, value=col)
    cell.font = fonte_cabecalho
    cell.fill = cor_cabecalho
    cell.alignment = alinhamento_centro
    cell.border = borda_fina
ws2.row_dimensions[3].height = 28

for r in range(4, 14):
    ws2.row_dimensions[r].height = 28
    for c in range(1, 7):
        cell = ws2.cell(row=r, column=c)
        cell.fill = cor_branco if r % 2 == 0 else cor_linha_alt
        cell.border = borda_fina
        cell.alignment = alinhamento_esq

for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 22 if col < 6 else 30

# ============================================================
# SHEET 3: AUTOCUIDADO
# ============================================================
ws3 = wb.create_sheet("Autocuidado")

ws3.merge_cells("A1:E1")
ws3["A1"] = "AUTOCUIDADO \u2014 PARA CUIDAR DE QUEM CUIDA"
ws3["A1"].font = fonte_titulo_play
ws3["A1"].alignment = alinhamento_centro
ws3.row_dimensions[1].height = 40

ws3.merge_cells("A2:E2")
ws3["A2"] = 'Voc\u00ea n\u00e3o pode servir de um copo vazio. Reserve tempo para recarregar.'
ws3["A2"].font = fonte_peq
ws3["A2"].alignment = alinhamento_centro
ws3.row_dimensions[2].height = 25

cabecalho3 = ["O que vou fazer?", "Quando?", "Dura\u00e7\u00e3o", "Como me senti antes", "Como me senti depois"]
for i, col in enumerate(cabecalho3, 1):
    cell = ws3.cell(row=4, column=i, value=col)
    cell.font = fonte_cabecalho
    cell.fill = cor_destaque
    cell.alignment = alinhamento_centro
    cell.border = borda_fina
ws3.row_dimensions[4].height = 28

sugestoes = [
    "Tomar um caf\u00e9 quente em sil\u00eancio",
    "Ler 10 p\u00e1ginas de um livro",
    "Fazer uma caminhada curta",
    "Tomar banho sem pressa",
    "Escrever no di\u00e1rio / journaling",
    "Ligar para uma amiga",
    "Assistir a um epis\u00f3dio da minha s\u00e9rie",
    "Fazer respira\u00e7\u00e3o consciente (5 min)",
    "Dan\u00e7ar uma m\u00fasica",
    "Fazer um skincare r\u00e1pido"
]

for i, sug in enumerate(sugestoes):
    r = 5 + i
    ws3.cell(row=r, column=1, value=sug).font = fonte_normal
    ws3.cell(row=r, column=1).alignment = alinhamento_esq
    ws3.cell(row=r, column=1).fill = cor_linha_alt if i % 2 == 0 else cor_branco
    for c in range(2, 6):
        ws3.cell(row=r, column=c).fill = cor_linha_alt if i % 2 == 0 else cor_branco
        ws3.cell(row=r, column=c).border = borda_fina
        ws3.cell(row=r, column=c).alignment = alinhamento_centro
    ws3.cell(row=r, column=1).border = borda_fina
    ws3.row_dimensions[r].height = 26

for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 22

# ============================================================
# SHEET 4: IMPREVISTOS E ADAPTACOES
# ============================================================
ws4 = wb.create_sheet("Imprevistos e Adapta\u00e7\u00f5es")

ws4.merge_cells("A1:D1")
ws4["A1"] = "IMPREVISTOS, CRISES E ADAPTA\u00c7\u00d5ES"
ws4["A1"].font = fonte_titulo_play
ws4["A1"].alignment = alinhamento_centro
ws4.row_dimensions[1].height = 40

ws4.merge_cells("A2:D2")
ws4["A2"] = "Registre o que saiu do planejamento e como voc\u00eas se adaptaram. Isso vira aprendizado para as pr\u00f3ximas semanas."
ws4["A2"].font = fonte_peq
ws4["A2"].alignment = alinhamento_centro
ws4.row_dimensions[2].height = 28

cabecalho4 = ["Dia / Hora", "O que aconteceu?", "Como lidamos?", "O que aprendi?"]
for i, col in enumerate(cabecalho4, 1):
    cell = ws4.cell(row=4, column=i, value=col)
    cell.font = fonte_cabecalho
    cell.fill = cor_imprev
    cell.alignment = alinhamento_centro
    cell.border = borda_fina
ws4.row_dimensions[4].height = 28

for r in range(5, 15):
    ws4.row_dimensions[r].height = 40
    for c in range(1, 5):
        cell = ws4.cell(row=r, column=c)
        cell.fill = cor_branco if r % 2 == 0 else cor_linha_alt
        cell.border = borda_fina
        cell.alignment = alinhamento_esq

for col in range(1, 5):
    ws4.column_dimensions[get_column_letter(col)].width = 18 if col == 1 else 35

# ============================================================
# SHEET 5: MEDICACOES
# ============================================================
ws5 = wb.create_sheet("Medica\u00e7\u00f5es e Suplementos")

ws5.merge_cells("A1:F1")
ws5["A1"] = "CONTROLE DE MEDICA\u00c7\u00d5ES E SUPLEMENTOS"
ws5["A1"].font = fonte_titulo_play
ws5["A1"].alignment = alinhamento_centro
ws5.row_dimensions[1].height = 40

cabecalho5 = ["Medica\u00e7\u00e3o", "Dosagem", "Hor\u00e1rios", "Dias da Semana", "Administrado?", "Observa\u00e7\u00f5es"]
for i, col in enumerate(cabecalho5, 1):
    cell = ws5.cell(row=3, column=i, value=col)
    cell.font = fonte_cabecalho
    cell.fill = cor_cabecalho
    cell.alignment = alinhamento_centro
    cell.border = borda_fina
ws5.row_dimensions[3].height = 28

for r in range(4, 10):
    ws5.row_dimensions[r].height = 28
    for c in range(1, 7):
        cell = ws5.cell(row=r, column=c)
        cell.fill = cor_branco if r % 2 == 0 else cor_linha_alt
        cell.border = borda_fina
        cell.alignment = alinhamento_centro

for col in range(1, 7):
    ws5.column_dimensions[get_column_letter(col)].width = 22

# ============================================================
# SALVAR
# ============================================================
caminho = "C:/Users/admin.suporte/Desktop/BLOG/planilha-semanal-mae-autista.xlsx"
wb.save(caminho)
print(f"Planilha salva em: {caminho}")
